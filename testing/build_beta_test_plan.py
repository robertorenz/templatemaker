#!/usr/bin/env python3
"""
Generates the Clarion Template Maker beta-test workbook (.xlsx).

    python testing/build_beta_test_plan.py

Output: testing/Clarion-Template-Maker-Beta-Test-Plan.xlsx
Re-run any time to regenerate; edit TEST_CASES below to add/adjust cases.
Requires: openpyxl  (pip install openpyxl)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

VERSION = "2.12.0"
OUT = os.path.join(os.path.dirname(__file__), "Clarion-Template-Maker-Beta-Test-Plan.xlsx")

# ---- palette (professional, no purple) -------------------------------------
NAVY    = "0E1A2B"
BLUE    = "0A66C2"
BLUE_LT = "EEF4FC"
GREY_LN = "D7DEE8"
ZEBRA   = "F5F8FC"
GREEN   = "D9F0E3"
RED     = "F9D7D2"
AMBER   = "FCEFCB"
GREYBG  = "ECEFF3"
WHITE   = "FFFFFF"
INK     = "22303F"

thin = Side(style="thin", color=GREY_LN)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)

RESULTS    = "Not Run,Pass,Fail,Blocked,Skipped"
SEVERITIES = "-,Critical,High,Medium,Low"
BUGSTATUS  = "Open,Fixed,Verified,Won't Fix,Duplicate"
TESTSTATUS = "Invited,Active,Completed,Dropped"


def header_row(ws, row, headers, widths):
    for c, (title, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 26


def banner(ws, ncols, title, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=16, color=WHITE)
    t.fill = PatternFill("solid", fgColor=BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(italic=True, size=10, color=INK)
    s.fill = PatternFill("solid", fgColor=BLUE_LT)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20


# ---- the test cases: (Area, Feature, Preconditions, Steps, Expected) --------
TEST_CASES = [
    # ---------------- Installation ----------------
    ("Installation", "Full installer runs",
     "ClarionTemplateToolsSetup.exe downloaded",
     "Double-click the setup .exe. If SmartScreen warns, choose More info > Run anyway. Step through the wizard and finish.",
     "Installer completes with no errors; a Start-menu shortcut and the install folder (templates, skills, docs) are created."),
    ("Installation", "Launch installed app",
     "Installed via the setup .exe",
     "Open the app from the Start menu shortcut.",
     "The Clarion Template Designer window opens; no error dialogs."),
    ("Installation", "Portable exe runs",
     "ClarionTemplateDesigner.exe (portable) downloaded",
     "Run the portable .exe directly from any folder (no install).",
     "App opens; the templates/skills/agents folders sit next to the exe and are usable."),
    ("Installation", "Version is correct",
     "App open",
     "Check the title bar / Help > About for the version.",
     f"Version shows {VERSION}."),
    ("Installation", "Uninstall is clean",
     "Installed via the setup .exe",
     "Uninstall from Settings > Apps (or the uninstaller).",
     "App and shortcuts removed; no leftover errors. (Re-install afterwards to keep testing.)"),

    # ---------------- Designer: files & documents ----------------
    ("Designer - Files", "Open a template",
     "App open",
     "File > Open and pick a shipped .tpl (e.g. templates\\myPie\\myPie.tpl).",
     "The prompt UI renders on the canvas at the real AT positions; image controls show their actual PNG icons."),
    ("Designer - Files", "Multiple documents / tabs",
     "App open",
     "Open two or three different .tpl files. Click between the document tabs at the top.",
     "Each template opens in its own tab; switching restores that document's selected part AND active tab."),
    ("Designer - Files", "Reopen refreshes canvas",
     "A template open",
     "Make an external change or reopen the same template.",
     "Opening a template refreshes the canvas immediately (no stale view)."),

    # ---------------- Designer: canvas editing ----------------
    ("Designer - Canvas", "Drag a control",
     "A template open with controls",
     "Drag a control to a new spot.",
     "The control moves; its AT x,y update; True layout mirrors the new position."),
    ("Designer - Canvas", "Resize a control",
     "A template open",
     "Grab a control's handle and resize it.",
     "The control resizes; AT width/height update accordingly."),
    ("Designer - Canvas", "Grid / ruler guides",
     "A template open",
     "Drag a guide off a ruler; snap a control to it; then drag the guide back onto the ruler.",
     "Controls snap to grid/guides; dragging a guide back to a ruler removes it."),
    ("Designer - Canvas", "Add controls",
     "A template open",
     "Use the Add command bar to insert each of: Label, String, Number, Spin, Check, Image, Group.",
     "Each new control appears with a fresh unique %symbol and is editable."),
    ("Designer - Canvas", "Add a new tab",
     "A template open",
     "Add a whole new #TAB via the Add bar.",
     "A new tab is created and selectable."),
    ("Designer - Canvas", "Reorder tabs by drag",
     "A template with 2+ tabs",
     "In the flow preview, drag one tab's header onto another.",
     "A caret shows where it will land; the whole #TAB...#ENDTAB block moves with it."),
    ("Designer - Canvas", "Group box parenting",
     "A template with a group box",
     "Drop a control inside a group box, then move the box.",
     "The control becomes a child; moving the box carries its contents."),
    ("Designer - Canvas", "Align / distribute / size",
     "A template with several controls",
     "Multi-select 3+ controls; use Arrange menu (or right-click) to align, distribute, and size together.",
     "Selected controls align/distribute/size as chosen."),
    ("Designer - Canvas", "Group / ungroup",
     "A template with several controls",
     "Select controls and press Ctrl+G; then Ctrl+Shift+G.",
     "Ctrl+G groups them into a box; Ctrl+Shift+G ungroups."),
    ("Designer - Canvas", "Smart alignment guides",
     "A template open",
     "Drag a control near another control's edges.",
     "Snap guides appear with a live spacing readout."),
    ("Designer - Canvas", "Copy / cut / paste / duplicate",
     "A template open",
     "Use Ctrl+C, Ctrl+X, Ctrl+V, Ctrl+D on a control.",
     "Controls copy/cut/paste/duplicate, each paste getting fresh %symbols."),
    ("Designer - Canvas", "Delete referenced symbol warns",
     "A control whose %symbol is used elsewhere",
     "Delete that control.",
     "A modal warning appears so you don't silently break code generation."),
    ("Designer - Canvas", "Visibility guides",
     "A template open",
     "Move a control partly off the window, outside its group box, or give a label too little room.",
     "The offending control/label is highlighted in red/amber."),

    # ---------------- Designer: properties & symbols ----------------
    ("Designer - Properties", "Symbol + Uses list",
     "A template open",
     "Select a control; look at the Properties pad.",
     "It shows the control's %symbol and a navigable Uses list of every reference; clicking an entry jumps to that line."),
    ("Designer - Properties", "Rename everywhere",
     "A control selected",
     "Click Rename and give a new symbol name.",
     "The symbol is renamed in the prompt AND in every reference at once."),
    ("Designer - Properties", "Prompt type / REQ / DEFAULT",
     "An added #PROMPT selected",
     "Edit its type, REQ, and DEFAULT in the editor.",
     "The prompt updates with the chosen type/REQ/DEFAULT."),
    ("Designer - Properties", "Tab rename / delete",
     "A template with tabs",
     "Right-click a tab header; rename it; delete another.",
     "Tab renames and deletes correctly."),
    ("Designer - Properties", "Tab WHERE() condition",
     "A template with tabs",
     "Right-click a tab; edit its WHERE(...) visibility condition.",
     "The WHERE() condition is saved on the tab."),
    ("Designer - Properties", "Font style flags",
     "A control selected",
     "Toggle bold / italic / underline and pick a color.",
     "Styles apply (PROP:FontStyle/FontColor); the IDE dialog font is shown read-only."),

    # ---------------- Designer: panels ----------------
    ("Designer - Panels", "Outline panel",
     "A template open",
     "Open the Outline panel; use its find box.",
     "It shows the #SHEET/#TAB/#BOXED/control tree; find filters it."),
    ("Designer - Panels", "Symbols panel",
     "A template open",
     "Open the Symbols panel; click a symbol.",
     "Lists every %symbol with a use count; clicking jumps to it."),
    ("Designer - Panels", "Problems panel",
     "A template open (ideally with an issue)",
     "Open the Problems panel; click an entry.",
     "Flags unbalanced blocks, duplicate/unused symbols, off-canvas/overlap, risky auto-prompts; clicking jumps to the spot."),

    # ---------------- Designer: source & save ----------------
    ("Designer - Source", "Find / replace",
     "Source panel visible",
     "Press Ctrl+F; find and replace text.",
     "Find/replace works in the source panel."),
    ("Designer - Source", "Autocomplete",
     "Source panel visible",
     "Start typing a %symbol and a #directive.",
     "Autocomplete suggests %symbols and #directives."),
    ("Designer - Source", "Preview changes diff",
     "Edited a template",
     "File > Preview changes.",
     "A colour-coded, per-file diff shows exactly what a save will write."),
    ("Designer - Source", "Save writes minimal change",
     "Edited control positions",
     "Save, then reopen the file.",
     "Only AT values change (plus any deletes/relocations); the rest of the file is untouched."),
    ("Designer - Source", "True layout matches",
     "A template open",
     "Toggle True layout preview; compare with the canvas and (if possible) with Clarion's AppGen.",
     "True layout mirrors the canvas and matches what Clarion draws."),

    # ---------------- Templates: registration ----------------
    ("Templates - Setup", "Register templates in Clarion",
     "Clarion 12 installed",
     "In the IDE: Setup > Template Registry > Register, and register each shipped .tpl.",
     "All templates register with no 'Expected an identifier'/parse errors."),

    # ---------------- Templates: each one ----------------
    ("Template: myPixel", "Diagnostic pixel",
     "myPixel.tpl registered",
     "Add myPixel to a window procedure; generate, compile, run.",
     "The per-window diagnostic pixel behaves as documented; app compiles and runs."),
    ("Template: showLine", "Where-am-I hotkey",
     "showLine.tpl registered",
     "Add showLine to a window; run; press Ctrl+Shift+P.",
     "Shows the current line/location as documented."),
    ("Template: identifier", "Procedure name hotkey",
     "identifier.tpl registered",
     "Add identifier to a window; run; press Ctrl+Shift+I.",
     "Displays the procedure name."),
    ("Template: myFuncs", "Global function library",
     "myFuncs registered",
     "Add the global myFuncs extension; call a provided function from code; compile and run.",
     "Functions are available globally and return correct results."),
    ("Template: myPie", "Pie chart + legend",
     "myPie registered; window has an IMAGE control",
     "Add myPieGlobal (app) + myPie (window); define segments; set legend/percent/depth/colors; generate, compile, run; resize the window.",
     "A pie with legend draws into the image; it redraws on open and resize; DO myPieRepaint updates it at run time."),
    ("Template: myFontChanger", "Global + per-list fonts",
     "myFontChanger registered",
     "Add the global font changer and a per-list font picker; compile, run.",
     "Fonts apply globally and per list as configured."),
    ("Template: myBackground", "Per-window background",
     "myBackground registered",
     "Add the global default + per-window extension; run; press Ctrl+Shift+B; pick a color, then an image; reopen the app.",
     "Background color/image apply per window and persist via INI across runs."),
    ("Template: myQR (online)", "QR from web service",
     "myQR registered; internet available; IMAGE control on window",
     "Add myQR; set a value (literal), size, ECC; enable auto-refresh; compile, run; scan with a phone; change the value and confirm refresh.",
     "A QR loads from the web service and scans to the value; auto-refresh redraws on change. (Needs internet.)"),
    ("Template: myQRDraw (window)", "Offline QR self-test",
     "myQRDrawGlobal + myQRDraw added; square IMAGE control",
     "Tick 'Draw self-test (HELLO WORLD)'; generate, compile, run; scan the code with a phone.",
     "An offline QR draws (no internet) and scans to exactly HELLO WORLD."),
    ("Template: myQRDraw (window)", "Offline QR live value",
     "myQRDraw on a window",
     "Untick self-test; set a real value (literal, then a variable + DO myQRDrawRepaint); resize the window.",
     "The code encodes the value, scans correctly, redraws on resize, and updates when the variable changes."),
    ("Template: myQRDrawReport", "QR on a report (per record)",
     "myQRDrawGlobal added; a REPORT with an IMAGE control (with a USE/field-equate) in the detail band",
     "Add myQRDrawReport; pick the report image control (the picker lists report controls); set a per-record value; print/preview; scan a row.",
     "One QR is drawn per record in the detail band and scans to that row's value."),
    ("Template: myQRDrawReport", "Report self-test",
     "myQRDrawReport on a report",
     "Tick the report self-test; print/preview; scan a printed code.",
     "Each row shows a QR reading HELLO WORLD."),

    # ---------------- QR encoder (technical) ----------------
    ("QR Encoder", "Automated tests (dev)",
     "Repo + .NET 9 SDK (technical testers only; else mark Skipped)",
     "Run: dotnet test designer/QrCodeCore.Tests",
     "All tests pass (round-trip via ZXing, Reed-Solomon ISO vector, HELLO WORLD golden matrix)."),

    # ---------------- General / UX ----------------
    ("General / UX", "Professional look, no purple",
     "App open",
     "Browse the UI and dialogs.",
     "Colors look professional; there is no purple anywhere."),
    ("General / UX", "Modal popups (not alerts)",
     "App open",
     "Trigger confirmations/warnings.",
     "They appear as modal popups, not plain alert message boxes."),
    ("General / UX", "DPI / resize",
     "App open",
     "Resize the window; try a high-DPI display / scaling if available.",
     "Layout scales cleanly; nothing is clipped or blurry."),
    ("General / UX", "Stability",
     "Whole session",
     "Use the app for an extended session across many templates.",
     "No crashes, hangs, or unhandled-exception dialogs."),
]

AREA_FILL = {}  # subtle per-area zebra handled below


def build():
    wb = Workbook()

    # =================== Read Me ===================
    ws = wb.active
    ws.title = "Read Me"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 110
    banner_cols = 3
    ws.merge_cells("A1:C1")
    c = ws["A1"]; c.value = f"Clarion Template Maker - Beta Test Plan (v{VERSION})"
    c.font = Font(bold=True, size=16, color=WHITE); c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(vertical="center", indent=1); ws.row_dimensions[1].height = 36

    lines = [
        ("", ""),
        ("Thank you for beta testing!", "h"),
        ("This workbook is your test script. Work through the Test Cases tab top to bottom, set a Result for each, "
         "and note anything odd. Log defects on the Bug Log tab. Add yourself on the Beta Testers tab.", "p"),
        ("What you are testing", "h"),
        ("1) The Template Designer app (install + the visual designer).", "p"),
        ("2) The Clarion templates it ships (register in Clarion 12, then generate / compile / run).", "p"),
        ("You do NOT need to test everything - do the areas assigned to you. Mark anything you can't run as 'Skipped'.", "p"),
        ("What you need", "h"),
        ("- Windows 10 or 11.", "p"),
        ("- For the template tests: Clarion 12 installed.", "p"),
        ("- A phone (any QR scanner / camera) for the QR tests.", "p"),
        ("- Internet only for the 'myQR (online)' test; everything else works offline.", "p"),
        ("How to record a result", "h"),
        ("In the Test Cases tab, use the Result drop-down: Not Run / Pass / Fail / Blocked / Skipped. "
         "Pass turns green, Fail turns red, Blocked turns amber. For a Fail, set a Severity, put your name + the date, "
         "and a short note (and a Bug ID if you logged one).", "p"),
        ("Severity guide", "h"),
        ("Critical = data loss / crash / can't proceed.  High = feature broken, no workaround.  "
         "Medium = broken but has a workaround.  Low = cosmetic / minor.", "p"),
        ("How to report bugs", "h"),
        ("Add a row on the Bug Log tab (give it a Bug ID like BUG-001 and reference the Test ID). "
         "GitHub issues: https://github.com/robertorenz/templatemaker/issues", "p"),
        ("The Summary tab tallies your results automatically.", "p"),
    ]
    r = 2
    for text, kind in lines:
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "h":
            cell.font = Font(bold=True, size=12, color=NAVY)
        elif kind == "p":
            cell.font = Font(size=11, color=INK); cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 30
        r += 1

    # =================== Beta Testers ===================
    bt = wb.create_sheet("Beta Testers")
    bt.sheet_view.showGridLines = False
    banner(bt, 8, "Beta Testers", "One row per tester. Fill in your details and which areas you'll cover.")
    bt_headers = ["Tester ID", "Name", "Email", "OS (10/11)", "Clarion ver.",
                  "Assigned areas", "Status", "Notes"]
    bt_widths  = [10, 22, 28, 12, 13, 30, 14, 30]
    header_row(bt, 4, bt_headers, bt_widths)
    bt.freeze_panes = "A5"
    example = ["BT-01", "Jane Example", "jane@example.com", "11", "12.0",
               "Designer + myQRDraw", "Active", "example row - overwrite me"]
    bt.cell(row=5, column=1)  # ensure row exists
    for i in range(5, 31):
        for c_ in range(1, 9):
            cell = bt.cell(row=i, column=c_, value=(example[c_-1] if i == 5 else None))
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            if i == 5:
                cell.font = Font(italic=True, color="7A8794")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
    dv_ts = DataValidation(type="list", formula1=f'"{TESTSTATUS}"', allow_blank=True)
    bt.add_data_validation(dv_ts); dv_ts.add(f"G5:G30")

    # =================== Test Cases ===================
    tc = wb.create_sheet("Test Cases")
    tc.sheet_view.showGridLines = False
    banner(tc, 11, "Test Cases",
           "Work top to bottom. Set Result via the drop-down; for a Fail add Severity, your name, date, note and a Bug ID.")
    headers = ["Test ID", "Area", "Feature", "Preconditions", "Steps",
               "Expected Result", "Result", "Severity", "Tester", "Date Tested", "Notes / Bug ID"]
    widths  = [9, 20, 22, 30, 46, 46, 11, 11, 16, 13, 30]
    header_row(tc, 4, headers, widths)
    tc.freeze_panes = "A5"

    start = 5
    # number IDs per area for readability: TC-001 ...
    for idx, (area, feature, pre, steps, expected) in enumerate(TEST_CASES):
        row = start + idx
        tcid = f"TC-{idx+1:03d}"
        values = [tcid, area, feature, pre, steps, expected, "Not Run", "", "", "", ""]
        for c_, val in enumerate(values, start=1):
            cell = tc.cell(row=row, column=c_, value=val)
            cell.border = BORDER
            if c_ in (1, 7, 8, 10):
                cell.alignment = CENTER
            else:
                cell.alignment = WRAP_TOP
            cell.font = Font(size=10, color=INK)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        tc.cell(row=row, column=2).font = Font(size=10, bold=True, color=NAVY)

    last = start + len(TEST_CASES) - 1
    tc.auto_filter.ref = f"A4:K{last}"

    # drop-downs
    dv_res = DataValidation(type="list", formula1=f'"{RESULTS}"', allow_blank=False)
    dv_sev = DataValidation(type="list", formula1=f'"{SEVERITIES}"', allow_blank=True)
    tc.add_data_validation(dv_res); dv_res.add(f"G{start}:G{last}")
    tc.add_data_validation(dv_sev); dv_sev.add(f"H{start}:H{last}")

    # conditional formatting on Result
    rng = f"G{start}:G{last}"
    tc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'],
        fill=PatternFill("solid", fgColor=GREEN), font=Font(color="11603F", bold=True)))
    tc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'],
        fill=PatternFill("solid", fgColor=RED), font=Font(color="8A2018", bold=True)))
    tc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'],
        fill=PatternFill("solid", fgColor=AMBER), font=Font(color="6F5210", bold=True)))
    tc.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Skipped"'],
        fill=PatternFill("solid", fgColor=GREYBG), font=Font(color="6B7785")))
    tc.print_options.horizontalCentered = True
    tc.page_setup.orientation = "landscape"
    tc.page_setup.fitToWidth = 1
    tc.page_setup.fitToHeight = 0
    tc.sheet_properties.pageSetUpPr.fitToPage = True

    # =================== Bug Log ===================
    bl = wb.create_sheet("Bug Log")
    bl.sheet_view.showGridLines = False
    banner(bl, 8, "Bug Log", "One row per defect. Reference the Test ID it came from.")
    bl_headers = ["Bug ID", "Test ID", "Summary", "Steps to reproduce / details",
                  "Severity", "Status", "Reported by", "Date"]
    bl_widths  = [9, 9, 34, 50, 11, 13, 16, 12]
    header_row(bl, 4, bl_headers, bl_widths)
    bl.freeze_panes = "A5"
    for i in range(5, 41):
        for c_ in range(1, 9):
            cell = bl.cell(row=i, column=c_)
            cell.border = BORDER; cell.alignment = WRAP_TOP; cell.font = Font(size=10, color=INK)
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
    dv_bsev = DataValidation(type="list", formula1=f'"{SEVERITIES}"', allow_blank=True)
    dv_bsta = DataValidation(type="list", formula1=f'"{BUGSTATUS}"', allow_blank=True)
    bl.add_data_validation(dv_bsev); dv_bsev.add("E5:E40")
    bl.add_data_validation(dv_bsta); dv_bsta.add("F5:F40")

    # =================== Summary ===================
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["A"].width = 3
    sm.column_dimensions["B"].width = 24
    sm.column_dimensions["C"].width = 14
    sm.merge_cells("A1:C1")
    c = sm["A1"]; c.value = "Results Summary"
    c.font = Font(bold=True, size=15, color=WHITE); c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(vertical="center", indent=1); sm.row_dimensions[1].height = 30

    col = f"'Test Cases'!G{start}:G{last}"
    rows = [
        ("Total test cases", f"=COUNTA('Test Cases'!A{start}:A{last})"),
        ("Pass",    f'=COUNTIF({col},"Pass")'),
        ("Fail",    f'=COUNTIF({col},"Fail")'),
        ("Blocked", f'=COUNTIF({col},"Blocked")'),
        ("Skipped", f'=COUNTIF({col},"Skipped")'),
        ("Not Run", f'=COUNTIF({col},"Not Run")'),
        ("Executed (Pass+Fail)", f'=COUNTIF({col},"Pass")+COUNTIF({col},"Fail")'),
        ("% Complete", f'=IFERROR((COUNTA(\'Test Cases\'!A{start}:A{last})-COUNTIF({col},"Not Run"))/COUNTA(\'Test Cases\'!A{start}:A{last}),0)'),
        ("Open bugs", f'=COUNTIF(\'Bug Log\'!F5:F40,"Open")'),
    ]
    r = 3
    for label, formula in rows:
        lc = sm.cell(row=r, column=2, value=label)
        lc.font = Font(bold=(label in ("Total test cases", "% Complete")), size=11, color=INK)
        lc.border = BORDER; lc.fill = PatternFill("solid", fgColor=BLUE_LT)
        vc = sm.cell(row=r, column=3, value=formula)
        vc.border = BORDER; vc.alignment = Alignment(horizontal="center")
        vc.font = Font(size=11, color=NAVY, bold=True)
        if label == "% Complete":
            vc.number_format = "0%"
        r += 1

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(TEST_CASES)} test cases)")


if __name__ == "__main__":
    build()
